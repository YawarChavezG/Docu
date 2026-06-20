"""
Tests para titulo_formulario_service: extraccion de titulos desde .xlsx y .docx.
Sesion 38 R3 — Servicios core.
"""
import io
import pytest

from app.services.titulo_formulario_service import (
    extraer_titulo_xlsx,
    extraer_titulo_docx,
    extraer_titulo,
)


# ════════════════════════════════════════════════════════════════
#  extraer_titulo_xlsx
# ════════════════════════════════════════════════════════════════


def _crear_xlsx(celda: str, valor: str) -> bytes:
    """Crea un .xlsx en memoria con un valor en una celda especifica (ej: 'C3')."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(celda)
    ws.cell(row=row, column=col, value=valor)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _crear_xlsx_con_filas(filas: list[list[str]]) -> bytes:
    """Crea un .xlsx con valores en filas (1-indexed)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, row_data in enumerate(filas, start=1):
        for j, val in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extraer_titulo_xlsx_celda_c3():
    """Titulo en C3 es el caso tipico de los formularios COFAR."""
    xlsx = _crear_xlsx("C3", "PROGRAMA DE EVALUACION DE PROVEEDORES")
    titulo = extraer_titulo_xlsx(xlsx)
    assert titulo == "PROGRAMA DE EVALUACION DE PROVEEDORES"


def test_extraer_titulo_xlsx_prioriza_fila_3_sobre_1():
    """Fila 3 tiene prioridad sobre fila 1 (estructura tipica COFAR)."""
    xlsx = _crear_xlsx_con_filas([
        ["ENCABEZADO DEL DEPARTAMENTO"],
        ["FORMULARIO: ACD-5-001"],
        ["TITULO REAL DEL DOCUMENTO"],
    ])
    titulo = extraer_titulo_xlsx(xlsx)
    assert titulo == "TITULO REAL DEL DOCUMENTO"


def test_extraer_titulo_xlsx_skip_keywords():
    """Celdas con keywords 'FORMULARIO', 'Versi', 'Vigente' se saltan."""
    xlsx = _crear_xlsx_con_filas([
        ["ENCABEZADO"],
        ["FORMULARIO: ABC-1-001"],
        ["Version: 01"],
        ["Vigente desde 2024"],
        ["REAL TITLE DEL DOCUMENTO"],
    ])
    titulo = extraer_titulo_xlsx(xlsx)
    assert titulo is not None
    assert "FORMULARIO" not in titulo
    assert "Version" not in titulo
    assert titulo == "REAL TITLE DEL DOCUMENTO"


def test_extraer_titulo_xlsx_sin_titulo_retorna_none():
    """Si no hay celdas con texto >10 chars, retorna None."""
    xlsx = _crear_xlsx_con_filas([
        ["A", "B", "C"],
        ["1", "2", "3"],
    ])
    assert extraer_titulo_xlsx(xlsx) is None


def test_extraer_titulo_xlsx_bytes_invalidos():
    """Bytes que no son un .xlsx valido: None (no exception)."""
    assert extraer_titulo_xlsx(b"no es un excel") is None


# ════════════════════════════════════════════════════════════════
#  extraer_titulo_docx
# ════════════════════════════════════════════════════════════════


def _crear_docx_con_parrafos(parrafos: list[str]) -> bytes:
    """Crea un .docx en memoria con los parrafos dados."""
    from docx import Document
    doc = Document()
    for p in parrafos:
        doc.add_paragraph(p)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def test_extraer_titulo_docx_primer_parrafo_significativo():
    """Primer parrafo con >10 chars y sin keywords es el titulo."""
    docx = _crear_docx_con_parrafos([
        "ESPECIFICACION DE ESTABILIDAD ACELERADA",
        "Aprobado por: Comite de Calidad",
    ])
    titulo = extraer_titulo_docx(docx)
    assert titulo == "ESPECIFICACION DE ESTABILIDAD ACELERADA"


def test_extraer_titulo_docx_skip_pagina():
    """Parrafos que contienen 'pagina' se saltan."""
    docx = _crear_docx_con_parrafos([
        "Pagina 1 de 10",
        "[Encabezado del documento]",
        "TITULO DEL DOCUMENTO",
    ])
    titulo = extraer_titulo_docx(docx)
    assert titulo is not None
    assert "pagina" not in titulo.lower()
    assert "TITULO DEL DOCUMENTO" in titulo


def test_extraer_titulo_docx_sin_parrafos_largos():
    """Si no hay parrafos >10 chars, retorna None."""
    docx = _crear_docx_con_parrafos(["A", "B", "C"])
    assert extraer_titulo_docx(docx) is None


# ════════════════════════════════════════════════════════════════
#  extraer_titulo (dispatcher principal)
# ════════════════════════════════════════════════════════════════


def test_extraer_titulo_xlsx_desde_dispatcher():
    """Dispatcher llama a extraer_titulo_xlsx para archivos .xlsx."""
    xlsx = _crear_xlsx("C3", "MI TITULO XLSX")
    titulo = extraer_titulo(xlsx, "formulario.xlsx")
    assert titulo == "MI TITULO XLSX"


def test_extraer_titulo_docx_desde_dispatcher():
    """Dispatcher llama a extraer_titulo_docx para archivos .docx."""
    docx = _crear_docx_con_parrafos(["MI TITULO DOCX"])
    titulo = extraer_titulo(docx, "documento.docx")
    assert titulo == "MI TITULO DOCX"


def test_extraer_titulo_doc_fallback_filename():
    """.doc legacy no se parsea, cae a fallback por nombre de archivo."""
    titulo = extraer_titulo(b"datos binarios .doc", "ACD-5-008-F01-MI FORMULARIO.doc")
    assert titulo is not None
    assert "FORMULARIO" in titulo or "MI" in titulo


def test_extraer_titulo_fallback_filename_sin_codigo():
    """Nombre de archivo sin codigo: se usa el nombre limpio."""
    titulo = extraer_titulo(b"", "mi_formulario_de_prueba.docx")
    assert titulo is not None
    assert "mi_formulario_de_prueba" in titulo
