"""
matriz_import_service.py — COFAR SGD (Sesion 8, tarea #12.2)

Carga masiva de roles y modulos desde el Excel de la matriz de abril
(`docs/Diagramas_Matrices/MATRICES/USUARIOS EXISTENTES A ABRIL.xlsx`,
hoja "LISTA PERSONAL").

Funciones principales:
- `parsear_excel(path)`: lee la hoja, normaliza, dedup por COFAR.
- `match_usuarios(rows, db)`: busca usuarios en BD por `ad_postal_code`.
- `aplicar_asignaciones(matched, db, update_existing)`: hace el bulk
  update (roles + modulos + flags) de forma idempotente.
- `run_import(...)`: orquesta las 3 funciones + dry-run support.

Reglas de negocio (ver docs/PR/MATRIZ-ABRIL-MAPEO.md):
- Match por `ad_postal_code == str(COFAR)`.
- Modulos UI se normalizan: "CONSULTAR DOCUMENTOS" -> "CONSULTAR_DOCUMENTOS"
  (replace " " por "_").
- Default: skip si el usuario ya tiene rol (no pisa decisiones manuales).
- Con `update_existing=True`: pisa rol, modulos, flags. NO toca `delegado_id`.
- N:M de roles/modulos: `ON CONFLICT DO NOTHING` (idempotente).
- `estado_delegacion`:
  - VISUALIZADOR -> `NA`
  - no-VISUALIZADOR -> `PENDIENTE` (delegado_id NO se asigna en este import).
- `delegado_id` NUNCA se toca (decision humana, protegida).
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Set, Tuple

import openpyxl
from sqlalchemy import delete, insert, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Modulo, Rol, Usuario
from app.models.usuario import EstadoDelegacion, usuario_modulos, usuario_roles

logger = logging.getLogger(__name__)

# ─── Constantes ───
HOJA = "LISTA PERSONAL"
HEADER_ESPERADO = [
    "COFAR", "NOMBRE", "POSICION", "ROL EN EL FLUJO", "MODULOS HABILITADOS",
    "\u00bfVISUALIZA O EXPORTA REPORTES?", "\u00bfREQUIERE DELEGADO?", "NOMBRE DELEGADO",
]

CODIGO_VISUALIZADOR = "VISUALIZADOR (CL-EVAL)"

CAMPOS_SI_NO: Dict[str, bool] = {"SI": True, "NO": False}

# Mapeo EXACTO texto del Excel -> CodigoModulo de la BD.
# No usar replace " " por "_" porque "BANDEJA DE TAREAS" -> "BANDEJA_TAREAS"
# (la particula "DE" se descarta, no se preserva).
MODULO_NORMALIZATION: Dict[str, str] = {
    "BANDEJA DE TAREAS": "BANDEJA_TAREAS",
    "MI BANDEJA": "MI_BANDEJA",
    "LISTA MAESTRA": "LISTA_MAESTRA",
    "CONSULTAR DOCUMENTOS": "CONSULTAR_DOCUMENTOS",
    "MIS EVALUACIONES": "MIS_EVALUACIONES",
    "ASISTENTE IA": "ASISTENTE_IA",
    "NUEVA SOLICITUD": "NUEVA_SOLICITUD",
    "PLANTILLAS DOCUMENTALES": "PLANTILLAS_DOCUMENTALES",
    "TODOS": "TODOS",
}

# Tamano del chunk para bulk operations (evita transacciones gigantes)
CHUNK_SIZE = 100


# ─── Dataclasses (resultados) ───
@dataclass
class MatrizRow:
    """Una fila del Excel ya normalizada."""
    cofar: str
    nombre: str
    posicion: str
    rol: str
    modulos: List[str]
    visualiza_reportes: bool
    requiere_delegado: bool


@dataclass
class MatchResult:
    """Resultado del match entre Excel y BD."""
    matched: List[Tuple[MatrizRow, int]] = field(default_factory=list)  # (row, user_id)
    unmatched: List[MatrizRow] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class ImportResult:
    """Resultado de aplicar las asignaciones a la BD."""
    roles_asignados: int = 0
    modulos_asignados: int = 0
    flags_actualizados: int = 0
    skipped_existing: int = 0
    warnings: List[str] = field(default_factory=list)
    errores: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errores


# ─── Helpers de normalizacion ───
def _normalizar_modulo(texto: str) -> str:
    """Lookup en el mapa MODULO_NORMALIZATION. None si no se reconoce."""
    return MODULO_NORMALIZATION.get(texto.strip().upper())


def _normalizar_rol(texto: str) -> str:
    return texto.strip().upper()


def _parsear_si_no(valor: object) -> bool:
    return CAMPOS_SI_NO.get(str(valor or "").strip().upper(), False)


def _cofar_a_str(cofar_raw: object) -> str:
    """Convierte el COFAR del Excel (int o str) a string de 8 digitos sin padding extra."""
    return str(int(cofar_raw))


# ─── 1. Parsear Excel ───
def parsear_excel(path: str | Path) -> Tuple[List[MatrizRow], List[str]]:
    """
    Lee la hoja LISTA PERSONAL del Excel de la matriz de abril.

    - Valida que el header coincida exactamente.
    - Ignora filas con COFAR vacio (filas vacias al final del Excel).
    - Dedup por COFAR (conserva la primera aparicion).
    - Cada modulo se normaliza con `_normalizar_modulo`.
    - Filas sin modulos habilitados se omiten con warning.

    Retorna (rows_validos, warnings).
    Lanza ValueError si la hoja o el header no existen.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel no encontrado: {path}")

    warnings: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if HOJA not in wb.sheetnames:
        raise ValueError(
            f"Hoja '{HOJA}' no encontrada en {path.name}. "
            f"Hojas disponibles: {wb.sheetnames}"
        )
    ws = wb[HOJA]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers != HEADER_ESPERADO:
        warnings.append(
            f"HEADER no coincide exactamente. "
            f"Esperado: {HEADER_ESPERADO}, leido: {headers}"
        )
    idx = {h: i for i, h in enumerate(headers)}

    cofar_vistos: Set[str] = set()
    rows: List[MatrizRow] = []

    for ridx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cofar_raw = fila[idx["COFAR"]]
        if cofar_raw is None:
            continue  # fila vacia al final del Excel

        cofar = _cofar_a_str(cofar_raw)
        if cofar in cofar_vistos:
            warnings.append(f"Fila {ridx}: COFAR {cofar} duplicado en el Excel, se ignora")
            continue
        cofar_vistos.add(cofar)

        rol_texto = fila[idx["ROL EN EL FLUJO"]] or ""
        rol = _normalizar_rol(rol_texto)
        if not rol:
            warnings.append(f"Fila {ridx} (COFAR {cofar}): sin ROL EN EL FLUJO, se ignora")
            continue

        modulos_texto = fila[idx["MODULOS HABILITADOS"]] or ""
        modulos_raw = [m.strip() for m in modulos_texto.split(",") if m.strip()]
        modulos: List[str] = []
        modulos_invalidos: List[str] = []
        for m in modulos_raw:
            codigo = _normalizar_modulo(m)
            if codigo is None:
                modulos_invalidos.append(m)
            else:
                modulos.append(codigo)
        if modulos_invalidos:
            warnings.append(
                f"Fila {ridx} (COFAR {cofar}): modulos no reconocidos "
                f"{modulos_invalidos}, se omiten (catalogo tiene {len(MODULO_NORMALIZATION)} entradas)"
            )
        if not modulos:
            warnings.append(f"Fila {ridx} (COFAR {cofar}): sin modulos validos, se ignora")
            continue

        rows.append(MatrizRow(
            cofar=cofar,
            nombre=str(fila[idx["NOMBRE"]] or "").strip(),
            posicion=str(fila[idx["POSICION"]] or "").strip(),
            rol=rol,
            modulos=modulos,
            visualiza_reportes=_parsear_si_no(fila[idx["\u00bfVISUALIZA O EXPORTA REPORTES?"]]),
            requiere_delegado=_parsear_si_no(fila[idx["\u00bfREQUIERE DELEGADO?"]]),
        ))

    return rows, warnings


# ─── 2. Matchear con BD ───
async def match_usuarios(rows: List[MatrizRow], db: AsyncSession) -> MatchResult:
    """
    Para cada row, busca el usuario en BD por `ad_postal_code == row.cofar`.

    Retorna MatchResult con listas separadas de matched/unmatched + warnings.
    Query unico a la BD (un solo SELECT con IN clause), no N+1.
    """
    result = MatchResult()
    cofars = list({r.cofar for r in rows})

    if not cofars:
        return result

    # Query masivo: trae todos los usuarios que matchean en un solo SELECT.
    # Orden por id ASC para preferir el usuario "real" (mas viejo) sobre el
    # stub del mismo COFAR (los stubs se crearon en el seed de DES, id alto).
    stmt = (
        select(Usuario.id, Usuario.ad_postal_code)
        .where(Usuario.ad_postal_code.in_(cofars))
        .order_by(Usuario.id.asc())
    )
    res = await db.execute(stmt)
    bd_index: Dict[str, int] = {}
    duplicados: Dict[str, List[int]] = {}
    for uid, cofar in res.all():
        if cofar in bd_index:
            duplicados.setdefault(cofar, [bd_index[cofar]]).append(uid)
        else:
            bd_index[cofar] = uid

    # Reportar duplicados como warning informativo
    for cofar, ids in duplicados.items():
        result.warnings.append(
            f"COFAR {cofar} tiene {len(ids) + 1} usuarios en BD (ids {ids}); "
            f"se prefiere id={bd_index[cofar]} (mas antiguo, probablemente real)"
        )

    for row in rows:
        user_id = bd_index.get(row.cofar)
        if user_id is not None:
            result.matched.append((row, user_id))
        else:
            result.unmatched.append(row)
            result.warnings.append(
                f"COFAR {row.cofar} ({row.nombre}): no existe en BD "
                f"(sin sincronizar desde AD). Skip."
            )

    return result


# ─── 3. Aplicar asignaciones ───
async def aplicar_asignaciones(
    matched: List[Tuple[MatrizRow, int]],
    db: AsyncSession,
    update_existing: bool = False,
) -> ImportResult:
    """
    Para cada (row, user_id) en `matched`, actualiza:
    - `roles` (N:M via `usuario_roles`)
    - `modulos` (N:M via `usuario_modulos`)
    - `requiere_delegado` (bool)
    - `visualiza_reportes` (bool)
    - `estado_delegacion` (`NA` | `PENDIENTE`)

    NO toca `delegado_id` (decision humana, protegida).
    NO toca campos de AD (email, cargo, area_id, etc.).

    Si `update_existing=False` (default), SKIP usuarios que ya tienen rol.
    Si `update_existing=True`, pisa todo (excepto `delegado_id`).
    """
    result = ImportResult()

    if not matched:
        return result

    # ─── Cache de catalogos (2 queries, una sola vez) ───
    r = await db.execute(select(Rol))
    roles_by_codigo: Dict[str, int] = {rol.codigo: rol.id for rol in r.scalars().all()}

    r = await db.execute(select(Modulo))
    modulos_by_codigo: Dict[str, int] = {m.codigo: m.id for m in r.scalars().all()}

    # ─── Validar catalogos antes de tocar la BD ───
    for row, _ in matched:
        if row.rol not in roles_by_codigo:
            result.errores.append(
                f"COFAR {row.cofar} ({row.nombre}): rol '{row.rol}' no esta en el catalogo"
            )
        for m in row.modulos:
            if m not in modulos_by_codigo:
                result.errores.append(
                    f"COFAR {row.cofar} ({row.nombre}): modulo '{m}' no esta en el catalogo"
                )
    if result.errores:
        return result

    # ─── Pre-cargar roles actuales de los usuarios a actualizar (evita N+1) ───
    user_ids = [uid for _, uid in matched]
    stmt = (
        select(Usuario.id, usuario_roles.c.rol_id)
        .select_from(Usuario)
        .outerjoin(usuario_roles, usuario_roles.c.usuario_id == Usuario.id)
        .where(Usuario.id.in_(user_ids))
    )
    res = await db.execute(stmt)
    roles_actuales_by_user: Dict[int, Set[int]] = {}
    for uid, rol_id in res.all():
        if rol_id is None:
            continue  # OUTER JOIN: usuario sin rol
        roles_actuales_by_user.setdefault(uid, set()).add(rol_id)

    # ─── Procesar en chunks (transacciones acotadas) ───
    for i in range(0, len(matched), CHUNK_SIZE):
        chunk = matched[i:i + CHUNK_SIZE]

        for row, user_id in chunk:
            roles_actuales = roles_actuales_by_user.get(user_id, set())

            # Skip si ya tiene rol y no estamos en modo update
            if roles_actuales and not update_existing:
                result.skipped_existing += 1
                result.warnings.append(
                    f"COFAR {row.cofar} ({row.nombre}): ya tiene rol asignado, "
                    f"skip (use --update-existing para sobrescribir)"
                )
                continue

            # 1) Actualizar roles (N:M): delete + insert atomico
            rol_id = roles_by_codigo[row.rol]
            await db.execute(
                delete(usuario_roles).where(usuario_roles.c.usuario_id == user_id)
            )
            await db.execute(insert(usuario_roles).values(usuario_id=user_id, rol_id=rol_id))
            result.roles_asignados += 1

            # 2) Actualizar modulos (N:M): delete + ON CONFLICT DO NOTHING
            modulos_ids = [modulos_by_codigo[m] for m in row.modulos]
            await db.execute(
                delete(usuario_modulos).where(usuario_modulos.c.usuario_id == user_id)
            )
            for mid in modulos_ids:
                stmt = pg_insert(usuario_modulos).values(usuario_id=user_id, modulo_id=mid)
                stmt = stmt.on_conflict_do_nothing(index_elements=["usuario_id", "modulo_id"])
                await db.execute(stmt)
            result.modulos_asignados += len(modulos_ids)

            # 3) Actualizar flags (bool + estado_delegacion)
            estado_deleg = (
                EstadoDelegacion.NA
                if row.rol == CODIGO_VISUALIZADOR
                else EstadoDelegacion.PENDIENTE
            )
            await db.execute(
                update(Usuario)
                .where(Usuario.id == user_id)
                .values(
                    requiere_delegado=row.requiere_delegado,
                    visualiza_reportes=row.visualiza_reportes,
                    estado_delegacion=estado_deleg,
                )
            )
            result.flags_actualizados += 1

        # Flush al final del chunk (no commit: el script CLI hace 1 commit al final)
        await db.flush()

    return result


# ─── 4. Orquestador (usado por el CLI y por tests) ───
async def run_import(
    db: AsyncSession,
    excel_path: str | Path,
    update_existing: bool = False,
) -> Tuple[MatchResult, ImportResult]:
    """
    Pipeline completo: parsear -> matchear -> aplicar.

    NO hace commit (lo hace el caller). Si hay errores de catalogo,
    no aplica nada (corta antes del paso 3).
    """
    rows, parse_warnings = parsear_excel(excel_path)
    logger.info(f"Excel parseado: {len(rows)} filas validas, {len(parse_warnings)} warnings")
    parse_warnings_str = [f"[parse] {w}" for w in parse_warnings]

    match_result = await match_usuarios(rows, db)
    match_result.warnings = parse_warnings_str + [
        f"[match] {w}" for w in match_result.warnings
    ]
    logger.info(
        f"Match: {len(match_result.matched)} matched, "
        f"{len(match_result.unmatched)} unmatched"
    )

    import_result = await aplicar_asignaciones(
        match_result.matched, db, update_existing=update_existing
    )
    return match_result, import_result
