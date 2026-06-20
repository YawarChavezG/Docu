"""
excel_export.py — Helper para generar exports XLSX profesionales con openpyxl.

Diseno:
  - Cabecera con fill verde COFAR pastel + texto blanco bold + altura 28
  - Freeze de la primera fila (encabezados siempre visibles al hacer scroll)
  - Auto-filtros en todas las columnas
  - Auto-width de columnas (calculado segun el contenido)
  - Bordes finos en toda la tabla
  - Fila de TOTALES al final con fill amarillo pastel + bold
  - Bandas alternadas (zebra striping) para lectura comoda
  - Alineacion: texto izquierda, numeros derecha, fechas centradas

Paleta pastel (alineada con el design system del frontend):
  - PRIMARY (cabecera):  #86EFAC  verde COFAR pastel
  - PRIMARY_TEXT:        #FFFFFF  blanco
  - TOTALES:             #FEF3C7  amarillo pastel
  - ZEBRA:               #F8FAFC  gris muy claro
  - BORDER:              #CBD5E1  slate-300
"""
import io
import logging
from datetime import date, datetime
from typing import Any, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ─── Paleta pastel ───
COLOR_PRIMARY = "86EFAC"        # verde COFAR pastel
COLOR_PRIMARY_TEXT = "FFFFFF"
COLOR_TOTALES = "FEF3C7"        # amarillo pastel
COLOR_TOTALES_TEXT = "78350F"   # amber-900
COLOR_ZEBRA = "F8FAFC"          # slate-50
COLOR_BORDER = "CBD5E1"         # slate-300
COLOR_TEXT = "1E293B"           # slate-800

# ─── Estilos reusables ───
_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=COLOR_PRIMARY_TEXT)
_FONT_CELL = Font(name="Calibri", size=10, color=COLOR_TEXT)
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color=COLOR_TOTALES_TEXT)
_FILL_HEADER = PatternFill("solid", fgColor=COLOR_PRIMARY)
_FILL_TOTAL = PatternFill("solid", fgColor=COLOR_TOTALES)
_FILL_ZEBRA = PatternFill("solid", fgColor=COLOR_ZEBRA)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_BORDER_THIN = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


def _autofit_columns(ws: Worksheet, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    """Calcula el ancho optimo de cada columna segun el contenido mas largo."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Considerar el header + el contenido mas largo
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                cell_val = row[col_idx - 1]
                if cell_val is None:
                    continue
                max_len = max(max_len, len(str(cell_val)))
        # Cap: min 10, max 60
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def build_excel(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    sheet_name: str = "Datos",
    title: Optional[str] = None,
    total_row: Optional[Sequence[Any]] = None,
    column_alignments: Optional[dict[int, str]] = None,
) -> bytes:
    """
    Genera un XLSX profesional en memoria y devuelve los bytes.

    Args:
        headers: Lista de cabeceras (1 por columna).
        rows: Filas de datos (cada fila es una secuencia con un valor por columna).
        sheet_name: Nombre de la hoja (max 31 chars, sin caracteres invalidos).
        title: Titulo opcional en la fila 1 (merge across all columns).
        total_row: Si se pasa, se agrega una fila "TOTAL" al final con este contenido.
                   La primera celda lleva la palabra "TOTAL", el resto los valores.
        column_alignments: Dict {1: 'left', 2: 'right', ...} para alinear columnas especificas.

    Returns:
        Bytes del XLSX listo para escribir en `response.body` o `StreamingResponse`.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita a 31 chars

    # ─── Title (opcional) ───
    current_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=COLOR_TEXT)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        current_row = 2  # Headers van en fila 2

    # ─── Headers ───
    header_row_idx = current_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_HEADER
        cell.border = _BORDER_THIN
    ws.row_dimensions[header_row_idx].height = 32
    current_row += 1

    # ─── Data rows ───
    for i, row in enumerate(rows):
        fill = _FILL_ZEBRA if i % 2 == 1 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = _FONT_CELL
            cell.border = _BORDER_THIN
            if fill is not None:
                cell.fill = fill
            # Alignment por defecto
            align_key = (column_alignments or {}).get(col_idx, "left")
            if align_key == "right":
                cell.alignment = _ALIGN_RIGHT
            elif align_key == "center":
                cell.alignment = _ALIGN_CENTER
            else:
                cell.alignment = _ALIGN_LEFT
        current_row += 1

    # ─── Total row (opcional) ───
    if total_row is not None:
        for col_idx, value in enumerate(total_row, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = _FONT_TOTAL
            cell.fill = _FILL_TOTAL
            cell.border = _BORDER_THIN
            cell.alignment = _ALIGN_CENTER if col_idx > 1 else _ALIGN_LEFT
        current_row += 1

    # ─── Auto-filtros en headers ───
    ws.auto_filter.ref = (
        f"A{header_row_idx}:{get_column_letter(len(headers))}{current_row - 1}"
    )

    # ─── Freeze: primera fila (headers) ───
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    # ─── Auto-width ───
    _autofit_columns(ws, headers, list(rows) + ([list(total_row)] if total_row else []))

    # ─── Serializar a bytes ───
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _to_csv_cell(value: Any) -> str:
    """Serializa un valor para CSV. None -> ''. Datetime -> ISO. Resto -> str."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "SI" if value else "NO"
    return str(value)


def build_csv(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    total_row: Optional[Sequence[Any]] = None,
) -> bytes:
    """
    Genera un CSV con BOM UTF-8 (asi Excel lo abre con acentos correctos).
    Separador: punto y coma (estandar hispanohablante para Excel).
    """
    import csv

    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_to_csv_cell(v) for v in row])
    if total_row is not None:
        writer.writerow([_to_csv_cell(v) for v in total_row])

    # BOM UTF-8 para que Excel detecte acentos
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
